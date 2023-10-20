"""This module is the controller of the Kniffel game. It should interact with the 
model where user input is utilized in accordance with the given data and with the view where the result of this interaction is turned into presentation."""

from model import *
from view import *
from setupscreen import *
import shelve
from openpyxl import load_workbook
import os

class Game:
    def __init__(self, playerlist=None, count=0):
        self.playerlist = playerlist
        if self.playerlist == None:
            self.playerlist = []
        self.count = count
        self.has_ended = False
        self.wb = None
        

    """def set_up_standard_scores(self):
        scores = Scoresheet()
        return scores"""

    """def set_up_user_determined_scores(self):
        pass"""
        
    def set_up_start_screen(self):
        """Start the first game screen to get player names, load a game or edit lists."""
        self.screen = Ui_MainWindow()
        f = QMainWindow()
        self.screen.setupUi(f, self)
        f.show()
        
    def set_up_game_screen(self):
        """Start the actual game screen."""
        self.screen = Ui_MainWindow()
        f = QMainWindow()
        self.screen.setupgameUi(f, self)
        f.show()
        
    def got_names(self, namelist):
        """Create Players from a given list of names. Start the game once players are created."""
        self.playerlist = list()
        self.namelist = namelist
        for index, i in enumerate(namelist):
            self.create_player(index, i)
        self.activeplayer = self.playerlist[0]
        self.set_up_game_screen()
        
    def start_round(self):
        """Set up tooltips for all players."""
        for i in self.playerlist:
            self.tooltip(i)
        
    def create_player(self, index, name):
        """Create a player with name and an index."""
        name = Player(name, index)
        self.playerlist.append(name)
        
    def create_autoplayer(self, index, name):
        name = Autokniffel(name, index)
        self.playerlist.append(name)
        
        
    def roll(self):
        """Roll all dice that are not marked as kept."""
        self.activeplayer.roll()
        self.display_take_options()
        return [x.value for x in self.activeplayer.mydice]
    
    def keep(self, die):
        """Mark individual die as kept."""
        self.activeplayer.keep(die)

    def end_turn(self):
        """Update tooltip, clear active player's dice and rollcount, switch to next player."""
        self.tooltip(self.activeplayer)
        self.next_player()
        
    def tooltip(self, player):
        """Create scoresheet as tooltip for a player."""
        tooltip = list()
        scores = Scoresheet.standard_scores()
        for key, value in scores.items():
            if key not in player.scored:
                tooltip.append(value[1] + "\t0\n")
            else:
                tooltip.append(player.scored[key][1] + "\t" + str(player.scored[key][0]) + "\n")
        toolst = "".join(tooltip)
        self.screen.update_tooltip(player.index, toolst)
        
    def take_option(self, option):
        """Change player's scores according to their pick (cross/take). Update total scores."""
        self.activeplayer.change_scored(option)
        self.activeplayer.total_top()
        self.activeplayer.total()

    def next_player(self):
        """Activate next player. Check if the game has ended."""
        self.activeplayer.clear()
        if self.activeplayer is self.playerlist[-1]:
            if len(self.activeplayer.scored) == len(Scoresheet.standard_scores()):
                self.has_ended = True
                self.end_scores()
                return
            self.playerlist[0].change_status()
            self.activeplayer = self.playerlist[0]
        else:
            self.activeplayer = self.playerlist[self.activeplayer.index + 1]
            self.activeplayer.change_status()
        try:
            self.screen.activate_player(self.activeplayer.index)
        except:
            pass

    def display_take_options(self):
        """Show player's scoring options."""
        self.screen.show_options(self.activeplayer.options_take())

    def display_cross_options(self):
        """Show player's cross-off options."""
        self.screen.show_options(self.activeplayer.options_cross(), 1)
        
    def end_scores(self):
        """Check for winner. If multiple players have highest score, there are multiple winners."""
        x = self.playerlist[0]
        winners = [x]
        for i in self.playerlist:
            if i > x:
                x = i
                winners.clear()
                winners.append(x)
            elif i == x and i not in winners:
                winners.append(i)
        self.screen.show_winners(winners)
        if self.sheetindex != None:
            self.get_list(self.sheetindex)

    def load_game(self, save=False):
        """Check for existing savegames and show a list of them, if existing. If User wants to save the game, the list will be shown as well."""
        files = list()
        for i in os.listdir():
            if i.endswith(".dat"):
                files.append(i[:-4])
        self.screen.show_load_files(files, save)
        
    def load_game_picked(self, name):
        """Get the name of the game that is supposed to be loaded. Open said file and restore the players accordingly."""
        savefile = shelve.open(name, "r")
        self.playerlist = savefile["playerlist"]
        for index, i in enumerate(self.playerlist):
            i = savefile[str(index)]
        for i in self.playerlist:
            if i.active:
                self.activeplayer = i
        self.screen.new_labels()
        for i in self.playerlist:
            self.tooltip(i)
        self.set_up_game_screen()
        savefile.close()


    def save_game(self):
        """User wants to save the game - get saved games to avoid multiple files with the same name."""
        self.load_game(save=True)
        
    def save_game_named(self, name):
        """Get a name input from user and check with files if name exists. If not, save the game and display success message."""
        try: 
            file = shelve.open(name, "r")
            file.close()
            self.screen.fileexist_popup(name)
        except:
            savefile = shelve.open(name)
            savefile["playerlist"] = self.playerlist
            for index, i in enumerate(self.playerlist):
                savefile[str(index)] = i
            savefile.close()
            self.screen.success(name)
        
    def quit_game(self):
        """Give user option to save game first, to go back or to simply quit."""
        sys.exit()
    
    def new_game(self):
        """Set up a new start screen."""
        self.set_up_start_screen()
    
    def create_list(self):
        """Show popup for name/scoring system input to create a list and if a game is already ongoing, show playernames."""
        if len(self.playerlist) > 0:
            names = [x.name for x in self.playerlist]
        else: names = None
        self.screen.name_popup(names)
            
    def from_list(self):
        """Enable User to start a game from an existing list. Names will automatically be the same."""
        self.view_list(newgame=True)
    
    
    def make_list(self, sheetname, names, scores):
        """Create an excel sheet with player names and scoring system."""
        wb = load_workbook("kniffellists.xlsx")
        sheet = wb.create_sheet(sheetname)
        sheet.append(scores)
        sheet.append(names)
        wb.save("kniffellists.xlsx")

    def view_list(self, newgame = False):
        """Let User choose from a list of lists which one they want to see."""
        wb = load_workbook("kniffellists.xlsx")
        sheets = wb.sheetnames
        wb.close()
        self.screen.get_list_name(sheets, newgame)
        
    def get_list(self, index, newgame=False):
        """Get an individual list from the excel workbook, according to index. Create a new list if intended."""
        wb = load_workbook("kniffellists.xlsx")
        sheet = wb.worksheets[index]
        letters = ["A", "B", "C", "D", "E", "F"]
        columns = list()
        if self.has_ended:
            self.add_end_scores(sheet)
            wb.save("kniffellists.xlsx")
        elif newgame:
            names = list()
            for i in letters:
                if sheet[i][1].value != None:
                    names.append(sheet[i][1].value)
            self.got_names(names)   
            self.sheetindex = index
            self.wb = wb
            self.set_up_game_screen()             
        else:    
            for i in letters:
                if sheet[i][1].value != None:
                    columns.append([x.value for x in sheet[i][1:]])
            self.screen.show_list(columns)
            

    def add_end_scores(self, sheet):
        """Add the final scores to a list. Automatically add points to existing points according to scoring system."""
        self.playerlist.sort(reverse=True)
        points = [x.value for x in sheet["1"]]
        lastrow = [x.value for x in sheet[sheet.max_row]]
        lastpoints = list()
        for i in lastrow:
            if i != None:
                i = i.split()
                lastpoints.append(int(i[0]))
        corrorder = list()
        for index, i in enumerate(self.playerlist):
            i.finscore = str(points[index]) + " " + str(i.scored[16][0])
        if "A" in points:
            for i in self.playerlist:
                if "X" not in [x[0] for x in i.scored.values()]:
                    i.finscore = i.finscore.split()
                    i.finscore[0] = int(i.finscore[0]) + 1
                    i.finscore[0] = str(i.finscore[0])
                    "".join(i.finscore)
        for i in [x.value for x in sheet["2"]]:
            for j in self.playerlist:
                if j.name == str(i):
                    corrorder.append(j.finscore)
        addedpoints = list()
        for index, i in enumerate(corrorder):
            full = i.split()
            point = int(full[0]) + lastpoints[index]
            addedpoints.append(str(point) + " " + full[1])
        sheet.append(addedpoints)